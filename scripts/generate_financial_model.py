#!/usr/bin/env python3
"""Generate Financial Model workbook.

Usage:
    python3 scripts/generate_financial_model.py --deal deals/<deal-folder>

Three sheets:
  Sheet 1 — P&L + SDE (copied from Financial Model workbook)
  Sheet 2 — Financial Model: SDE weighting, financing with deal costs, DSCR
  Sheet 3 — DSCR Sensitivity: rate×price and SDE×price matrices with conditional formatting

Circular reference (SBA guarantee fee depends on SBA loan which depends on TPC
which depends on deal costs which includes guarantee fee) is resolved algebraically:
  TPC = [PP×(1 - GF_PCT×SN_PCT) + BaseCosts] / [1 - GF_PCT×(1 - BE_PCT)]
"""

import argparse
import glob
from copy import copy
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
RESULT_FONT = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
RESULT_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
THICK_BORDER = Border(
    top=Side(style="medium", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100", bold=True)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)

DOLLAR_FMT = '$#,##0'
PCT_FMT = '0.0%'
DSCR_FMT = '0.00"x"'
MULT_FMT = '0.00"x"'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_section(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).font = SECTION_FONT
        ws.cell(row=row, column=c).fill = SECTION_FILL


def style_total(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THICK_BORDER


def style_result(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = RESULT_FONT
        cell.fill = RESULT_FILL
        cell.border = THICK_BORDER


def inp(ws, r, c, val, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = INPUT_FILL
    if fmt:
        cell.number_format = fmt
    return cell


def fml(ws, r, c, formula, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c, value=formula)
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = TOTAL_FONT
    return cell


def lbl(ws, r, c, text, indent=False):
    cell = ws.cell(row=r, column=c, value=("  " + text) if indent else text)
    if indent:
        cell.font = Font(name="Calibri", size=9, color="888888")
    return cell


def helper_font(ws, r, c):
    ws.cell(row=r, column=c).font = Font(italic=True, size=9, color="666666")
    ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Financial Model
# ─────────────────────────────────────────────────────────────────────────────

def build_calculator(wb, deal_name="", moderate_sde_row=None, pnl_years=None):
    ws = wb.active
    ws.title = "Financial Model"
    ws.sheet_properties.tabColor = "1F4E79"
    set_col_widths(ws, [42, 16, 10, 16, 3, 38, 16])

    refs = {}

    # Title
    ws.cell(row=1, column=1, value=f"FINANCIAL MODEL — {deal_name}")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14)
    ws.cell(row=2, column=1, value="Yellow cells are editable inputs — all calculations update automatically")
    ws.cell(row=2, column=1).font = Font(name="Calibri", italic=True, size=10, color="666666")

    # ═══ SDE WEIGHTED AVERAGE ═══════════════════════════════════════════════
    r = 4
    ws.cell(row=r, column=1, value="SDE WEIGHTED AVERAGE")
    style_section(ws, r, 1, 4)

    r = 5
    for c, h in enumerate(["Period", "SDE", "Weight", "Weighted SDE"], 1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 1, 4)

    # Build periods from the P&L + SDE tab's actual years and moderate SDE row
    pnl = "'P&L + SDE'"
    sde_row = moderate_sde_row or 59

    if pnl_years and len(pnl_years) >= 2:
        n = len(pnl_years)
        periods = []
        for i, (label, col) in enumerate(pnl_years):
            col_letter = get_column_letter(col)
            sde_ref = f"={pnl}!{col_letter}{sde_row}"
            if n == 2:
                weight = 0.40 if i == 0 else 0.60
            elif n == 3:
                weight = [0.15, 0.35, 0.50][i]
            elif n == 4:
                weight = [0.00, 0.25, 0.35, 0.40][i]
            else:
                weight = round(1.0 / n, 2) if i < n - 1 else round(1.0 - round(1.0 / n, 2) * (n - 1), 2)
            periods.append((label, sde_ref, weight))
    else:
        periods = [
            ("Year 1", f"={pnl}!B{sde_row}", 0.40),
            ("Year 2", f"={pnl}!C{sde_row}", 0.60),
        ]

    sde_start = 6
    for i, (label, sde_ref, weight) in enumerate(periods):
        r = sde_start + i
        ws.cell(row=r, column=1, value=label)
        fml(ws, r, 2, sde_ref, DOLLAR_FMT)
        inp(ws, r, 3, weight, PCT_FMT)
        fml(ws, r, 4, f"=B{r}*C{r}", DOLLAR_FMT)
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = THIN_BORDER
    sde_end = sde_start + len(periods) - 1

    r = sde_end + 1  # 10
    lbl(ws, r, 1, "Weight Check (must = 100%)", indent=True)
    fml(ws, r, 3, f"=SUM(C{sde_start}:C{sde_end})", PCT_FMT)

    r = sde_end + 2  # 11
    ws.cell(row=r, column=1, value="WEIGHTED AVERAGE SDE")
    fml(ws, r, 2, f"=SUM(D{sde_start}:D{sde_end})", DOLLAR_FMT, bold=True)
    style_total(ws, r, 1, 4)
    refs['weighted_sde_row'] = r  # 11

    # ═══ SDE ADJUSTMENTS ════════════════════════════════════════════════════
    r += 2  # 13
    ws.cell(row=r, column=1, value="SDE ADJUSTMENTS")
    style_section(ws, r, 1, 4)

    r += 1  # 14
    ws.cell(row=r, column=1, value="Weighted Average SDE")
    fml(ws, r, 2, f"=B{refs['weighted_sde_row']}", DOLLAR_FMT)
    refs['sde_used_row'] = r

    r += 1  # 15
    ws.cell(row=r, column=1, value="Owner Replacement Cost (annual)")
    inp(ws, r, 2, 100000, DOLLAR_FMT)
    refs['replacement_row'] = r

    r += 1  # 16
    ws.cell(row=r, column=1, value="ROBS Annual Ongoing Cost")
    inp(ws, r, 2, 0, DOLLAR_FMT)
    refs['robs_ongoing_row'] = r

    r += 1  # 17
    ws.cell(row=r, column=1, value="EFFECTIVE SDE (for DSCR)")
    fml(ws, r, 2,
        f"=B{refs['sde_used_row']}-B{refs['replacement_row']}-B{refs['robs_ongoing_row']}",
        DOLLAR_FMT, bold=True)
    style_total(ws, r, 1, 4)
    refs['effective_sde_row'] = r  # 17

    # ═══ PURCHASE & FINANCING ═══════════════════════════════════════════════
    r += 2  # 19
    ws.cell(row=r, column=1, value="PURCHASE & FINANCING")
    style_section(ws, r, 1, 4)

    r += 1  # 20
    ws.cell(row=r, column=1, value="Purchase Price")
    inp(ws, r, 2, 5500000, DOLLAR_FMT)
    refs['price_row'] = r

    r += 1  # 21
    ws.cell(row=r, column=1, value="Asking Price Multiple (on Weighted SDE)")
    fml(ws, r, 2, f"=B{refs['price_row']}/B{refs['sde_used_row']}", MULT_FMT)

    r += 1  # 22
    ws.cell(row=r, column=1, value="Seller Note % (of purchase price)")
    inp(ws, r, 2, 0.10, PCT_FMT)
    refs['sn_pct_row'] = r

    r += 1  # 23
    ws.cell(row=r, column=1, value="Buyer Equity % (of total project cost)")
    inp(ws, r, 2, 0.10, PCT_FMT)
    refs['be_pct_row'] = r

    r += 1  # 24
    ws.cell(row=r, column=1, value="SBA Guarantee Fee Rate (est)")
    inp(ws, r, 2, 0.03, PCT_FMT)
    refs['gf_pct_row'] = r

    r += 2  # 26 — TPC (formula filled after deal costs panel defines base_costs_row)
    ws.cell(row=r, column=1, value="Total Project Cost (PP + all deal costs)")
    refs['tpc_row'] = r

    r += 1  # 27
    ws.cell(row=r, column=1, value="Seller Note Amount (% × Purchase Price)")
    fml(ws, r, 2, f"=ROUND(B{refs['sn_pct_row']}*B{refs['price_row']},0)", DOLLAR_FMT)
    refs['sn_row'] = r

    r += 1  # 28
    ws.cell(row=r, column=1, value="Buyer Equity Amount (% × Total Project Cost)")
    fml(ws, r, 2, f"=ROUND(B{refs['be_pct_row']}*B{refs['tpc_row']},0)", DOLLAR_FMT)
    refs['be_row'] = r

    r += 1  # 29
    ws.cell(row=r, column=1, value="SBA Loan Amount (remainder)")
    fml(ws, r, 2, f"=B{refs['tpc_row']}-B{refs['sn_row']}-B{refs['be_row']}", DOLLAR_FMT, bold=True)
    refs['sba_row'] = r

    r += 1  # 30
    lbl(ws, r, 1, "SBA as % of Total Project Cost", indent=True)
    fml(ws, r, 2, f"=B{refs['sba_row']}/B{refs['tpc_row']}", PCT_FMT)

    r += 1  # 31
    lbl(ws, r, 1, "Check: SN + BE + SBA = TPC", indent=True)
    fml(ws, r, 2, f"=B{refs['sn_row']}+B{refs['be_row']}+B{refs['sba_row']}", DOLLAR_FMT)
    style_total(ws, r, 1, 2)

    # ═══ LOAN TERMS ═════════════════════════════════════════════════════════
    r += 2  # 33
    ws.cell(row=r, column=1, value="LOAN TERMS")
    style_section(ws, r, 1, 4)

    r += 1  # 34
    ws.cell(row=r, column=1, value="SBA Interest Rate (annual)")
    inp(ws, r, 2, 0.105, PCT_FMT)
    refs['sba_rate_row'] = r

    r += 1  # 35
    ws.cell(row=r, column=1, value="SBA Loan Term (years)")
    inp(ws, r, 2, 10, '0')
    refs['sba_term_row'] = r

    r += 1  # 36
    ws.cell(row=r, column=1, value="Seller Note Interest Rate")
    inp(ws, r, 2, 0.06, PCT_FMT)
    refs['seller_rate_row'] = r

    r += 1  # 37
    ws.cell(row=r, column=1, value="Seller Note Amortization (years)")
    inp(ws, r, 2, 10, '0')
    refs['seller_amort_row'] = r

    r += 1  # 38
    ws.cell(row=r, column=1, value="Seller Note Standby Period (years)")
    inp(ws, r, 2, 0, '0')
    refs['seller_standby_row'] = r

    r += 1  # 39
    ws.cell(row=r, column=1, value="Seller Note Maturity (year from closing)")
    inp(ws, r, 2, 5, '0')
    lbl(ws, r, 4, "Balloon due this year. Set = standby + amort for no balloon", indent=True)
    refs['maturity_row'] = r

    # ═══ DEBT SERVICE ═══════════════════════════════════════════════════════
    r += 2  # 41
    ws.cell(row=r, column=1, value="DEBT SERVICE")
    style_section(ws, r, 1, 4)

    r += 1  # 42
    ws.cell(row=r, column=1, value="SBA Monthly Payment")
    fml(ws, r, 2,
        f"=IF(B{refs['sba_rate_row']}=0,B{refs['sba_row']}/(B{refs['sba_term_row']}*12),"
        f"-PMT(B{refs['sba_rate_row']}/12,B{refs['sba_term_row']}*12,B{refs['sba_row']}))",
        DOLLAR_FMT)
    refs['sba_monthly_row'] = r

    r += 1  # 43
    ws.cell(row=r, column=1, value="SBA Annual Debt Service")
    fml(ws, r, 2, f"=B{refs['sba_monthly_row']}*12", DOLLAR_FMT, bold=True)
    refs['sba_annual_row'] = r

    r += 1  # 44
    ws.cell(row=r, column=1, value="Seller Note Monthly Payment (post-standby)")
    fml(ws, r, 2,
        f"=IF(B{refs['seller_rate_row']}=0,B{refs['sn_row']}/(B{refs['seller_amort_row']}*12),"
        f"-PMT(B{refs['seller_rate_row']}/12,B{refs['seller_amort_row']}*12,B{refs['sn_row']}))",
        DOLLAR_FMT)
    refs['seller_monthly_row'] = r

    r += 1  # 45
    ws.cell(row=r, column=1, value="Seller Note Annual Debt Service")
    fml(ws, r, 2, f"=B{refs['seller_monthly_row']}*12", DOLLAR_FMT, bold=True)
    refs['seller_annual_row'] = r

    r += 1  # 46
    ws.cell(row=r, column=1, value="Seller Note Balloon Payment")
    # Payment years = maturity - standby. If payment years >= amortization, no balloon.
    # Remaining principal = -FV(rate/12, payment_years*12, -monthly_pmt, principal)
    fml(ws, r, 2,
        f"=IF((B{refs['maturity_row']}-B{refs['seller_standby_row']})>=B{refs['seller_amort_row']},0,"
        f"-FV(B{refs['seller_rate_row']}/12,"
        f"(B{refs['maturity_row']}-B{refs['seller_standby_row']})*12,"
        f"-B{refs['seller_monthly_row']},"
        f"B{refs['sn_row']}))",
        DOLLAR_FMT)
    refs['balloon_row'] = r

    r += 1  # 47
    lbl(ws, r, 1, "Balloon due (year from closing)", indent=True)
    fml(ws, r, 2,
        f"=IF(B{refs['balloon_row']}=0,\"N/A — fully amortized\","
        f"\"Year \"&B{refs['maturity_row']})")

    r += 2  # 49
    ws.cell(row=r, column=1, value="Total DS (During Standby)")
    fml(ws, r, 2,
        f"=IF(B{refs['seller_standby_row']}=0,"
        f"B{refs['sba_annual_row']}+B{refs['seller_annual_row']},"
        f"B{refs['sba_annual_row']})",
        DOLLAR_FMT, bold=True)
    style_total(ws, r, 1, 2)
    refs['ds_standby_row'] = r

    r += 1  # 50
    ws.cell(row=r, column=1, value="Total DS (Post-Standby — SBA + Seller)")
    fml(ws, r, 2, f"=B{refs['sba_annual_row']}+B{refs['seller_annual_row']}", DOLLAR_FMT, bold=True)
    style_total(ws, r, 1, 2)
    refs['ds_full_row'] = r

    r += 1  # 51
    ws.cell(row=r, column=1, value="Total DS (Balloon Year — SBA + Seller + Balloon)")
    fml(ws, r, 2,
        f"=B{refs['sba_annual_row']}+B{refs['seller_annual_row']}+B{refs['balloon_row']}",
        DOLLAR_FMT, bold=True)
    style_total(ws, r, 1, 2)
    refs['ds_balloon_row'] = r

    # ═══ DSCR ANALYSIS ══════════════════════════════════════════════════════
    r += 2  # 53
    ws.cell(row=r, column=1, value="DSCR ANALYSIS")
    style_section(ws, r, 1, 4)

    r += 1  # 54
    ws.cell(row=r, column=1, value="Minimum Required DSCR")
    inp(ws, r, 2, 1.25, DSCR_FMT)
    refs['min_dscr_row'] = r

    r += 2  # 56
    ws.cell(row=r, column=1, value="DSCR (During Standby)")
    fml(ws, r, 2, f"=B{refs['effective_sde_row']}/B{refs['ds_standby_row']}", DSCR_FMT)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=14)
    refs['dscr_standby_row'] = r

    r += 1  # 57
    lbl(ws, r, 1, "Pass/Fail", indent=True)
    fml(ws, r, 2, f'=IF(B{refs["dscr_standby_row"]}>=B{refs["min_dscr_row"]},"PASS","FAIL")')

    r += 2  # 59
    ws.cell(row=r, column=1, value="DSCR (Post-Standby)")
    fml(ws, r, 2, f"=B{refs['effective_sde_row']}/B{refs['ds_full_row']}", DSCR_FMT)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=14)
    refs['dscr_full_row'] = r

    r += 1  # 60
    lbl(ws, r, 1, "Pass/Fail", indent=True)
    fml(ws, r, 2, f'=IF(B{refs["dscr_full_row"]}>=B{refs["min_dscr_row"]},"PASS","FAIL")')

    r += 2  # 62
    ws.cell(row=r, column=1, value="DSCR (Balloon Year)")
    fml(ws, r, 2,
        f"=IF(B{refs['balloon_row']}=0,B{refs['dscr_full_row']},"
        f"B{refs['effective_sde_row']}/B{refs['ds_balloon_row']})",
        DSCR_FMT)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=14)
    refs['dscr_balloon_row'] = r

    r += 1  # 63
    lbl(ws, r, 1, "Pass/Fail", indent=True)
    fml(ws, r, 2,
        f'=IF(B{refs["balloon_row"]}=0,"N/A",'
        f'IF(B{refs["dscr_balloon_row"]}>=B{refs["min_dscr_row"]},"PASS","FAIL"))')

    r += 2  # 65
    ws.cell(row=r, column=1, value="VERDICT")
    fml(ws, r, 2,
        f'=IF(AND(B{refs["dscr_full_row"]}>=B{refs["min_dscr_row"]},'
        f'OR(B{refs["balloon_row"]}=0,B{refs["dscr_balloon_row"]}>=B{refs["min_dscr_row"]})),"FEASIBLE",'
        f'IF(B{refs["dscr_standby_row"]}>=B{refs["min_dscr_row"]},'
        f'"MARGINAL — standby only","DOES NOT PENCIL"))')
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=13)

    # ═══ RIGHT PANEL: DEAL COSTS ════════════════════════════════════════════
    fc, gc = 6, 7

    cr = 4
    ws.cell(row=cr, column=fc, value="DEAL COSTS (rolled into SBA financing)")
    style_section(ws, cr, fc, gc)

    cr = 5
    ws.cell(row=cr, column=fc, value="Cost Item")
    ws.cell(row=cr, column=gc, value="Amount")
    style_header(ws, cr, fc, gc)

    costs = [
        ("Attorney / Legal", 20000),
        ("Quality of Earnings (QofE)", 17000),
        ("Accounting / Tax Advisory", 0),
        ("Business Valuation", 0),
        ("Environmental / Inspection", 0),
        ("ROBS Setup Cost", 0),
        ("Working Capital Reserve", 100000),
        ("Other Closing Costs", 0),
    ]
    cost_start = 6
    for i, (label, val) in enumerate(costs):
        cr = cost_start + i
        ws.cell(row=cr, column=fc, value=label)
        inp(ws, cr, gc, val, DOLLAR_FMT)
        ws.cell(row=cr, column=fc).border = THIN_BORDER
        ws.cell(row=cr, column=gc).border = THIN_BORDER
    cost_end = cost_start + len(costs) - 1  # 13

    cr = cost_end + 1  # 14
    ws.cell(row=cr, column=fc, value="BASE DEAL COSTS (before guarantee fee)")
    fml(ws, cr, gc, f"=SUM(G{cost_start}:G{cost_end})", DOLLAR_FMT, bold=True)
    style_total(ws, cr, fc, gc)
    refs['base_costs_row'] = cr  # 14

    cr += 1  # 15
    ws.cell(row=cr, column=fc, value="SBA Guarantee Fee (estimated)")
    fml(ws, cr, gc,
        f"=B{refs['tpc_row']}-B{refs['price_row']}-G{refs['base_costs_row']}", DOLLAR_FMT)
    refs['gf_amount_row'] = cr

    cr += 1  # 16
    ws.cell(row=cr, column=fc, value="TOTAL DEAL COSTS")
    fml(ws, cr, gc, f"=G{refs['base_costs_row']}+G{refs['gf_amount_row']}", DOLLAR_FMT, bold=True)
    style_total(ws, cr, fc, gc)
    refs['total_costs_row'] = cr

    # ── Fill in TPC formula (algebraic resolution of circular ref) ──
    # TPC = [PP×(1 - GF_PCT×SN_PCT) + BaseCosts] / [1 - GF_PCT×(1 - BE_PCT)]
    ws.cell(row=refs['tpc_row'], column=2, value=(
        f"=(B{refs['price_row']}*(1-B{refs['gf_pct_row']}*B{refs['sn_pct_row']})"
        f"+G{refs['base_costs_row']})"
        f"/(1-B{refs['gf_pct_row']}*(1-B{refs['be_pct_row']}))"
    ))
    ws.cell(row=refs['tpc_row'], column=2).number_format = DOLLAR_FMT
    style_total(ws, refs['tpc_row'], 1, 2)

    # ═══ RIGHT PANEL: CASH TO CLOSE ═════════════════════════════════════════
    cr += 2  # 18
    ws.cell(row=cr, column=fc, value="CASH TO CLOSE")
    style_section(ws, cr, fc, gc)

    cr += 1  # 19
    ws.cell(row=cr, column=fc, value="Buyer Equity (down payment)")
    fml(ws, cr, gc, f"=B{refs['be_row']}", DOLLAR_FMT)
    cash_eq_row = cr

    cr += 1  # 20
    ws.cell(row=cr, column=fc, value="Less: ROBS 401(k) Used for Equity")
    inp(ws, cr, gc, 0, DOLLAR_FMT)
    refs['robs_401k_row'] = cr

    cr += 1  # 21
    ws.cell(row=cr, column=fc, value="NET OUT-OF-POCKET CASH")
    fml(ws, cr, gc, f"=G{cash_eq_row}-G{refs['robs_401k_row']}", DOLLAR_FMT)
    style_result(ws, cr, fc, gc)

    ws.freeze_panes = "A4"
    return refs


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: DSCR Sensitivity
# ─────────────────────────────────────────────────────────────────────────────

def build_sensitivity(wb, refs):
    ws = wb.create_sheet("DSCR Sensitivity")
    ws.sheet_properties.tabColor = "C00000"
    set_col_widths(ws, [24, 14, 14, 14, 14, 14])

    dc = "'Financial Model'"
    price_offsets = [-0.20, -0.10, 0, 0.10, 0.20]
    price_labels = ["-20%", "-10%", "Base", "+10%", "+20%"]

    # Title
    ws.cell(row=1, column=1, value="DSCR SENSITIVITY ANALYSIS")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14)
    ws.cell(row=2, column=1, value="Post-standby DSCR — green >= 1.25x (PASS), red < 1.25x (FAIL)")
    ws.cell(row=2, column=1).font = Font(name="Calibri", italic=True, size=10, color="666666")

    # ═══ MATRIX 1: RATE × PRICE ═════════════════════════════════════════════
    r = 4
    ws.cell(row=r, column=1, value="INTEREST RATE × PURCHASE PRICE")
    style_section(ws, r, 1, 6)

    # Header
    r = 5
    ws.cell(row=r, column=1, value="SBA Rate ↓  /  Price →")
    for j, label in enumerate(price_labels):
        ws.cell(row=r, column=j + 2, value=label)
    style_header(ws, r, 1, 6)

    # Helper row: Purchase prices
    r = 6
    ws.cell(row=r, column=1, value="Purchase Price")
    helper_font(ws, r, 1)
    for j, pct in enumerate(price_offsets):
        c = j + 2
        fml(ws, r, c, f"={dc}!B{refs['price_row']}*(1+{pct})", '$#,##0')
        helper_font(ws, r, c)
    price_helper_row = r

    # Helper row: TPC for each price
    # TPC = (PP*(1-GF*SN) + BC) / (1 - GF*(1-BE))
    r = 7
    ws.cell(row=r, column=1, value="Total Project Cost")
    helper_font(ws, r, 1)
    for j in range(len(price_offsets)):
        c = j + 2
        cl = get_column_letter(c)
        pp = f"{cl}{price_helper_row}"
        tpc_f = (f"=({pp}*(1-{dc}!B{refs['gf_pct_row']}*{dc}!B{refs['sn_pct_row']})"
                 f"+{dc}!G{refs['base_costs_row']})"
                 f"/(1-{dc}!B{refs['gf_pct_row']}*(1-{dc}!B{refs['be_pct_row']}))")
        fml(ws, r, c, tpc_f, '$#,##0')
        helper_font(ws, r, c)
    tpc_helper_row = r

    # Helper row: SBA loan for each price
    # SBA = TPC*(1-BE_PCT) - SN_PCT*PP
    r = 8
    ws.cell(row=r, column=1, value="SBA Loan")
    helper_font(ws, r, 1)
    for j in range(len(price_offsets)):
        c = j + 2
        cl = get_column_letter(c)
        fml(ws, r, c,
            f"={cl}{tpc_helper_row}*(1-{dc}!B{refs['be_pct_row']})"
            f"-{dc}!B{refs['sn_pct_row']}*{cl}{price_helper_row}",
            '$#,##0')
        helper_font(ws, r, c)
    sba_helper_row = r

    # Helper row: Seller annual DS for each price (fixed seller rate)
    r = 9
    ws.cell(row=r, column=1, value="Seller Annual DS")
    helper_font(ws, r, 1)
    for j in range(len(price_offsets)):
        c = j + 2
        cl = get_column_letter(c)
        sn = f"({dc}!B{refs['sn_pct_row']}*{cl}{price_helper_row})"
        fml(ws, r, c,
            f"=-PMT({dc}!B{refs['seller_rate_row']}/12,"
            f"{dc}!B{refs['seller_amort_row']}*12,{sn})*12",
            '$#,##0')
        helper_font(ws, r, c)
    seller_ds_helper_row = r

    # Helper row: Total DS at base rate (used by Matrix 2)
    r = 10
    ws.cell(row=r, column=1, value="Total DS (base rate)")
    helper_font(ws, r, 1)
    for j in range(len(price_offsets)):
        c = j + 2
        cl = get_column_letter(c)
        fml(ws, r, c,
            f"=-PMT({dc}!B{refs['sba_rate_row']}/12,"
            f"{dc}!B{refs['sba_term_row']}*12,{cl}{sba_helper_row})*12"
            f"+{cl}{seller_ds_helper_row}",
            '$#,##0')
        helper_font(ws, r, c)
    total_ds_base_row = r

    # Rate rows
    rates = [0.085, 0.090, 0.095, 0.100, 0.105, 0.110, 0.115, 0.120]
    rate_start = 12
    for i, rate in enumerate(rates):
        r = rate_start + i
        ws.cell(row=r, column=1, value=rate)
        ws.cell(row=r, column=1).number_format = '0.0%'
        ws.cell(row=r, column=1).font = Font(bold=True)
        for j in range(len(price_offsets)):
            c = j + 2
            cl = get_column_letter(c)
            # DSCR = eff_sde / (-PMT(rate/12, term*12, SBA)*12 + Seller_DS)
            fml(ws, r, c,
                f"={dc}!B{refs['effective_sde_row']}"
                f"/(-PMT($A{r}/12,{dc}!B{refs['sba_term_row']}*12,{cl}${sba_helper_row})*12"
                f"+{cl}${seller_ds_helper_row})",
                DSCR_FMT)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    rate_end = rate_start + len(rates) - 1

    # Conditional formatting for Matrix 1
    rng1 = f"B{rate_start}:F{rate_end}"
    ws.conditional_formatting.add(rng1,
        CellIsRule(operator='greaterThanOrEqual', formula=['1.25'],
                   fill=GREEN_FILL, font=GREEN_FONT))
    ws.conditional_formatting.add(rng1,
        CellIsRule(operator='lessThan', formula=['1.25'],
                   fill=RED_FILL, font=RED_FONT))

    # ═══ MATRIX 2: SDE × PRICE ══════════════════════════════════════════════
    r = rate_end + 3
    ws.cell(row=r, column=1, value="SDE SCENARIO × PURCHASE PRICE")
    style_section(ws, r, 1, 6)
    r += 1
    ws.cell(row=r, column=1,
            value="At current interest rate and deal terms from Financial Model")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="666666")

    r += 1
    ws.cell(row=r, column=1, value="SDE ↓  /  Price →")
    for j, label in enumerate(price_labels):
        ws.cell(row=r, column=j + 2, value=label)
    style_header(ws, r, 1, 6)
    sde_header = r

    # SDE scenario rows
    sde_scenarios = [
        ("SDE -30%", -0.30),
        ("SDE -20%", -0.20),
        ("SDE -10%", -0.10),
        ("Base SDE", 0),
        ("SDE +10%", 0.10),
        ("SDE +20%", 0.20),
    ]

    sde_start = sde_header + 1
    for i, (label, pct) in enumerate(sde_scenarios):
        r = sde_start + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = Font(bold=True)

        # varied_eff_sde = weighted_sde*(1+pct) - replacement - robs_ongoing
        eff_sde_expr = (
            f"({dc}!B{refs['sde_used_row']}*(1+{pct})"
            f"-{dc}!B{refs['replacement_row']}"
            f"-{dc}!B{refs['robs_ongoing_row']})"
        )

        for j in range(len(price_offsets)):
            c = j + 2
            cl = get_column_letter(c)
            # DSCR = varied_eff_sde / total_DS_at_base_rate
            fml(ws, r, c, f"={eff_sde_expr}/{cl}${total_ds_base_row}", DSCR_FMT)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    sde_end = sde_start + len(sde_scenarios) - 1

    # Conditional formatting for Matrix 2
    rng2 = f"B{sde_start}:F{sde_end}"
    ws.conditional_formatting.add(rng2,
        CellIsRule(operator='greaterThanOrEqual', formula=['1.25'],
                   fill=GREEN_FILL, font=GREEN_FONT))
    ws.conditional_formatting.add(rng2,
        CellIsRule(operator='lessThan', formula=['1.25'],
                   fill=RED_FILL, font=RED_FONT))

    ws.freeze_panes = "A4"


def find_moderate_sde_row(ws):
    """Scan the P&L + SDE sheet for the Moderate SDE row."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and "moderate" in cell.value.lower():
            return cell.row
    return None


def find_pnl_years(ws):
    """Find the year columns from the P&L + SDE header row (row 4)."""
    years = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and str(val).strip().isdigit():
            years.append((str(val).strip(), col))
        elif val and isinstance(val, str) and val.strip().startswith("TTM"):
            years.append((val.strip(), col))
    return years


def copy_pnl_sheet(source_path, target_wb):
    """Copy the 'P&L + SDE' sheet and return (moderate_sde_row, years)."""
    src_wb = load_workbook(source_path)
    src_ws = src_wb["P&L + SDE"]

    moderate_row = find_moderate_sde_row(src_ws)
    years = find_pnl_years(src_ws)

    dst_ws = target_wb.create_sheet("P&L + SDE", 0)
    dst_ws.sheet_properties.tabColor = "2F5233"

    for col_idx, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_idx].width = dim.width

    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row,
                                max_col=src_ws.max_column):
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column,
                                   value=cell.value)
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.fill = copy(cell.fill)
                dst_cell.border = copy(cell.border)
                dst_cell.alignment = copy(cell.alignment)
                dst_cell.number_format = cell.number_format

    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    src_wb.close()
    return moderate_row, years


def main():
    parser = argparse.ArgumentParser(description="Generate Financial Model workbook")
    parser.add_argument("--deal", required=True, help="Path to deal folder (e.g. deals/my-deal)")
    args = parser.parse_args()

    deal_dir = Path(args.deal)
    if not deal_dir.is_dir():
        print(f"Error: {deal_dir} is not a directory")
        return

    # Derive deal name from folder (convert slug to title case)
    deal_name = deal_dir.name.replace("-", " ").title()

    # Find the Financial Model workbook
    fm_candidates = glob.glob(str(deal_dir / "*Financial Model*.xlsx"))
    if not fm_candidates:
        print(f"Error: no Financial Model workbook found in {deal_dir}")
        return
    pnl_path = fm_candidates[0]

    output = deal_dir / f"{deal_name} - Financial Model.xlsx"

    wb = Workbook()
    moderate_row, pnl_years = copy_pnl_sheet(pnl_path, wb)
    if moderate_row:
        print(f"Found Moderate SDE at row {moderate_row}")
    else:
        print("Warning: could not find Moderate SDE row — defaulting to row 59")
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    dc_ws = wb.create_sheet("Financial Model")
    wb.active = wb.sheetnames.index("Financial Model")
    refs = build_calculator(wb, deal_name, moderate_row, pnl_years)
    build_sensitivity(wb, refs)
    wb.save(str(output))
    print(f"Saved: {output}")


if __name__ == "__main__":
    main()
