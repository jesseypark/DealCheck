#!/usr/bin/env python3
"""Generate a P&L + SDE workbook from deal_state.json.

Usage:
    python3 scripts/generate_pnl_workbook.py --deal deals/<deal-folder>

Produces a single-sheet .xlsx (P&L + SDE) with:
  - Revenue breakdown, COGS, Gross Profit
  - Operating expenses with all line items
  - Net Income
  - SDE add-backs grouped by classification (Verified / Plausible / Disputed) with subtotals
  - Three SDE tiers: Conservative, Moderate, Aggressive
  - Key metrics

The Financial Model script (generate_financial_model.py) copies this sheet and adds
the Financial Model + DSCR Sensitivity tabs.
"""

import argparse
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styling ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D6E4C8", end_color="D6E4C8", fill_type="solid")
SDE_FILL = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
SDE_FONT_WHITE = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color="666666")
NOTE_FONT = Font(name="Calibri", size=9, color="666666")
FAINT_FONT = Font(name="Calibri", italic=True, size=9, color="888888")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
THICK_BORDER_TOP = Border(
    top=Side(style="medium", color="2F5233"),
    bottom=Side(style="medium", color="2F5233"),
)
ACCT_FMT = '#,##0'
ACCT_FMT_NEG = '#,##0;[Red](#,##0)'
PCT_FMT = '0.0%'


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_section_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).font = SECTION_FONT
        ws.cell(row=row, column=c).fill = SECTION_FILL


def style_total_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THICK_BORDER_TOP


def style_sde_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SDE_FONT_WHITE
        cell.fill = SDE_FILL
        cell.border = THICK_BORDER_TOP


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def thin_border_row(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


# ── Data extraction ──────────────────────────────────────────────────────────

def _find_year_entry(entries, year):
    """Find a multi-value entry matching a given year."""
    if not entries:
        return None
    for entry in entries:
        val = entry.get("value", {})
        if isinstance(val, dict) and val.get("year") == year:
            return val
    return None


def _slug_to_label(slug):
    """Convert snake_case key to Title Case label."""
    return slug.replace("_", " ").title()


def load_deal_data(deal_dir):
    """Load and structure data from deal_state.json for spreadsheet generation."""
    ds_path = Path(deal_dir) / "deal_state.json"
    with open(ds_path) as f:
        ds = json.load(f)

    meta = ds.get("metadata", {})
    dim2 = ds.get("dimensions", {}).get("2_financial_performance", {})
    sde_recon = ds.get("analysis", {}).get("sde_reconstruction", {})

    deal_name = meta.get("deal_name", Path(deal_dir).name.replace("-", " ").title())

    # Determine which years we have data for
    revenue_entries = dim2.get("revenue_by_year", [])
    years = sorted(set(
        entry["value"]["year"]
        for entry in revenue_entries
        if isinstance(entry.get("value", {}), dict)
        and isinstance(entry["value"].get("year"), int)
    ))

    # Determine confidence basis
    sources = set()
    for entry in revenue_entries:
        src = entry.get("source", "").lower()
        if "tax" in src:
            sources.add("tax")
        elif "p&l" in src or "profit" in src or "quickbooks" in src:
            sources.add("pl")
        elif "cim" in src:
            sources.add("cim")
    if "tax" in sources:
        confidence_label = "Tax Return Basis (Confidence 0.70)"
    elif "pl" in sources:
        confidence_label = "P&L Basis (Confidence 0.50)"
    else:
        confidence_label = "CIM Basis (Confidence 0.50)"

    # Revenue breakdown
    revenue_data = {}
    for yr in years:
        entry = _find_year_entry(revenue_entries, yr)
        if entry:
            revenue_data[yr] = {
                "total": entry.get("amount", 0),
                "breakdown": entry.get("breakdown", {}),
            }

    # COGS breakdown
    cogs_entries = dim2.get("cogs_by_year", [])
    cogs_data = {}
    for yr in years:
        entry = _find_year_entry(cogs_entries, yr)
        if entry:
            cogs_data[yr] = {
                "total": entry.get("amount", 0),
                "breakdown": entry.get("breakdown", {}),
            }

    # OpEx breakdown
    opex_entries = dim2.get("opex_breakdown", [])
    opex_data = {}
    for yr in years:
        entry = _find_year_entry(opex_entries, yr)
        if entry:
            opex_data[yr] = entry.get("categories", {})

    # Net income
    ni_entries = dim2.get("net_income_by_year", [])
    ni_data = {}
    for yr in years:
        entry = _find_year_entry(ni_entries, yr)
        if entry:
            ni_data[yr] = entry.get("amount", 0)

    # SDE add-backs by year — split into verified / plausible / disputed
    addbacks_by_year = {}
    for yr in years:
        yr_str = str(yr)
        if yr_str not in sde_recon:
            continue
        yr_data = sde_recon[yr_str]

        verified = yr_data.get("addbacks_verified", [])
        # addbacks_disputed contains both yellow (plausible) and red (disputed)
        mixed = yr_data.get("addbacks_disputed", [])
        plausible = [a for a in mixed if a.get("rating") == "yellow"]
        disputed = [a for a in mixed if a.get("rating") == "red"]

        addbacks_by_year[yr] = {
            "base_net_income": yr_data.get("base_net_income", ni_data.get(yr, 0)),
            "verified": verified,
            "plausible": plausible,
            "disputed": disputed,
            "sde_conservative": yr_data.get("sde_conservative"),
            "sde_moderate": yr_data.get("sde_moderate"),
            "sde_aggressive": yr_data.get("sde_aggressive"),
        }

    return {
        "deal_name": deal_name,
        "confidence_label": confidence_label,
        "years": years,
        "revenue": revenue_data,
        "cogs": cogs_data,
        "opex": opex_data,
        "net_income": ni_data,
        "addbacks": addbacks_by_year,
    }


# ── Spreadsheet generation ───────────────────────────────────────────────────

def build_pnl_sheet(wb, data):
    ws = wb.active
    ws.title = "P&L + SDE"
    ws.sheet_properties.tabColor = "2F5233"

    years = data["years"]
    num_years = len(years)
    # Columns: A=label, B..=years, spacer, YoY%, Rating, Notes
    year_cols = list(range(2, 2 + num_years))  # B, C, ...
    spacer_col = 2 + num_years  # after last year
    yoy_col = spacer_col + 1
    rating_col = yoy_col + 1
    notes_col = rating_col + 1
    max_col = notes_col

    widths = [42] + [16] * num_years + [3, 12, 12, 40]
    set_col_widths(ws, widths)

    # ── Title ──
    row = 1
    ws.cell(row=row, column=1, value=data["deal_name"])
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14)
    row = 2
    ws.cell(row=row, column=1, value=f"P&L Recast + SDE Reconstruction — {data['confidence_label']}")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT

    # ── Column headers ──
    row = 4
    ws.cell(row=row, column=1, value="")
    for i, yr in enumerate(years):
        ws.cell(row=row, column=year_cols[i], value=str(yr))
    ws.cell(row=row, column=yoy_col, value="YoY %")
    ws.cell(row=row, column=notes_col, value="Notes")
    style_header_row(ws, row, max_col)

    # ── REVENUE ──
    row = 5
    ws.cell(row=row, column=1, value="REVENUE")
    style_section_row(ws, row, max_col)

    # Collect all revenue line item keys across all years
    rev_keys = []
    seen = set()
    for yr in years:
        for key in data["revenue"].get(yr, {}).get("breakdown", {}).keys():
            if key not in seen:
                rev_keys.append(key)
                seen.add(key)

    rev_start_row = row + 1
    for key in rev_keys:
        row += 1
        ws.cell(row=row, column=1, value=_slug_to_label(key))
        for i, yr in enumerate(years):
            val = data["revenue"].get(yr, {}).get("breakdown", {}).get(key, 0)
            cell = ws.cell(row=row, column=year_cols[i], value=val)
            cell.number_format = ACCT_FMT
        thin_border_row(ws, row, 1, max_col)
    rev_end_row = row

    # Total Revenue
    row += 1
    ws.cell(row=row, column=1, value="TOTAL REVENUE")
    rev_row = row
    for i, yr in enumerate(years):
        c = year_cols[i]
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f"=SUM({cl}{rev_start_row}:{cl}{rev_end_row})"
        ws.cell(row=row, column=c).number_format = ACCT_FMT
    if num_years >= 2:
        prev_cl = get_column_letter(year_cols[-2])
        last_cl = get_column_letter(year_cols[-1])
        ws.cell(row=row, column=yoy_col).value = f"=({last_cl}{row}-{prev_cl}{row})/{prev_cl}{row}"
        ws.cell(row=row, column=yoy_col).number_format = PCT_FMT
    style_total_row(ws, row, max_col)

    # ── COGS ──
    has_cogs = any(data["cogs"].get(yr, {}).get("total", 0) != 0 for yr in years)
    if has_cogs:
        row += 2
        ws.cell(row=row, column=1, value="COST OF GOODS SOLD")
        style_section_row(ws, row, max_col)

        cogs_keys = []
        seen = set()
        for yr in years:
            for key in data["cogs"].get(yr, {}).get("breakdown", {}).keys():
                if key not in seen:
                    cogs_keys.append(key)
                    seen.add(key)

        cogs_start_row = row + 1
        for key in cogs_keys:
            row += 1
            ws.cell(row=row, column=1, value=_slug_to_label(key))
            for i, yr in enumerate(years):
                val = data["cogs"].get(yr, {}).get("breakdown", {}).get(key, 0)
                cell = ws.cell(row=row, column=year_cols[i], value=val)
                cell.number_format = ACCT_FMT
            thin_border_row(ws, row, 1, max_col)
        cogs_end_row = row

        row += 1
        ws.cell(row=row, column=1, value="TOTAL COGS")
        cogs_row = row
        for i in range(num_years):
            c = year_cols[i]
            cl = get_column_letter(c)
            ws.cell(row=row, column=c).value = f"=SUM({cl}{cogs_start_row}:{cl}{cogs_end_row})"
            ws.cell(row=row, column=c).number_format = ACCT_FMT
        style_total_row(ws, row, max_col)

        # Gross Profit
        row += 1
        ws.cell(row=row, column=1, value="GROSS PROFIT")
        gp_row = row
        for i in range(num_years):
            c = year_cols[i]
            cl = get_column_letter(c)
            ws.cell(row=row, column=c).value = f"={cl}{rev_row}-{cl}{cogs_row}"
            ws.cell(row=row, column=c).number_format = ACCT_FMT
        if num_years >= 2:
            prev_cl = get_column_letter(year_cols[-2])
            last_cl = get_column_letter(year_cols[-1])
            ws.cell(row=row, column=yoy_col).value = f"=({last_cl}{row}-{prev_cl}{row})/{prev_cl}{row}"
            ws.cell(row=row, column=yoy_col).number_format = PCT_FMT
        style_total_row(ws, row, max_col)

        row += 1
        ws.cell(row=row, column=1, value="  Gross Margin %")
        ws.cell(row=row, column=1).font = FAINT_FONT
        for i in range(num_years):
            c = year_cols[i]
            cl = get_column_letter(c)
            ws.cell(row=row, column=c).value = f"={cl}{gp_row}/{cl}{rev_row}"
            ws.cell(row=row, column=c).number_format = PCT_FMT
    else:
        gp_row = rev_row

    # ── OPERATING EXPENSES ──
    row += 2
    ws.cell(row=row, column=1, value="OPERATING EXPENSES")
    style_section_row(ws, row, max_col)

    opex_keys = []
    seen = set()
    for yr in years:
        for key in data["opex"].get(yr, {}).keys():
            if key not in seen:
                opex_keys.append(key)
                seen.add(key)

    expense_start_row = row + 1
    for key in opex_keys:
        row += 1
        ws.cell(row=row, column=1, value=_slug_to_label(key))
        for i, yr in enumerate(years):
            val = data["opex"].get(yr, {}).get(key, 0)
            cell = ws.cell(row=row, column=year_cols[i], value=val if val else "")
            if val:
                cell.number_format = ACCT_FMT
        # YoY for items present in both last two years
        if num_years >= 2:
            val_prev = data["opex"].get(years[-2], {}).get(key, 0)
            val_last = data["opex"].get(years[-1], {}).get(key, 0)
            if val_prev and val_last:
                prev_cl = get_column_letter(year_cols[-2])
                last_cl = get_column_letter(year_cols[-1])
                ws.cell(row=row, column=yoy_col).value = f"=({last_cl}{row}-{prev_cl}{row})/{prev_cl}{row}"
                ws.cell(row=row, column=yoy_col).number_format = PCT_FMT
        # Note absent items
        for i, yr in enumerate(years):
            val = data["opex"].get(yr, {}).get(key, 0)
            if not val and any(data["opex"].get(y, {}).get(key, 0) for y in years):
                if i == num_years - 1:
                    ws.cell(row=row, column=notes_col, value=f"Absent in {yr}")
                    ws.cell(row=row, column=notes_col).font = NOTE_FONT
        thin_border_row(ws, row, 1, max_col)
    expense_end_row = row

    # Total Expenses
    row += 1
    ws.cell(row=row, column=1, value="TOTAL OPERATING EXPENSES")
    total_exp_row = row
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f"=SUM({cl}{expense_start_row}:{cl}{expense_end_row})"
        ws.cell(row=row, column=c).number_format = ACCT_FMT
    if num_years >= 2:
        prev_cl = get_column_letter(year_cols[-2])
        last_cl = get_column_letter(year_cols[-1])
        ws.cell(row=row, column=yoy_col).value = f"=({last_cl}{row}-{prev_cl}{row})/{prev_cl}{row}"
        ws.cell(row=row, column=yoy_col).number_format = PCT_FMT
    style_total_row(ws, row, max_col)

    # ── NET INCOME ──
    row += 2
    ws.cell(row=row, column=1, value="NET INCOME")
    ni_row = row
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f"={cl}{gp_row}-{cl}{total_exp_row}"
        ws.cell(row=row, column=c).number_format = ACCT_FMT_NEG
    if num_years >= 2:
        prev_cl = get_column_letter(year_cols[-2])
        last_cl = get_column_letter(year_cols[-1])
        ws.cell(row=row, column=yoy_col).value = f"=({last_cl}{row}-{prev_cl}{row})/{prev_cl}{row}"
        ws.cell(row=row, column=yoy_col).number_format = PCT_FMT
    style_total_row(ws, row, max_col)

    # ── SDE ADD-BACKS ──
    row += 2
    ws.cell(row=row, column=1, value="SDE ADD-BACKS")
    style_section_row(ws, row, max_col)

    row += 1
    ws.cell(row=row, column=1, value="")
    for i, yr in enumerate(years):
        ws.cell(row=row, column=year_cols[i], value=str(yr))
    ws.cell(row=row, column=rating_col, value="Rating")
    ws.cell(row=row, column=notes_col, value="Notes")
    style_header_row(ws, row, max_col)

    # Build unified add-back item lists across years
    # Each group: list of (item_name, {year: amount}, rating_label, note)
    def _build_group(group_key, rating_label):
        items_by_name = {}
        for yr in years:
            ab = data["addbacks"].get(yr, {})
            for item in ab.get(group_key, []):
                name = item["item"]
                if name not in items_by_name:
                    items_by_name[name] = {
                        "amounts": {},
                        "rating": rating_label,
                        "note": item.get("reasoning", ""),
                    }
                items_by_name[name]["amounts"][yr] = item.get("amount", 0)
        return list(items_by_name.items())

    verified_items = _build_group("verified", "VERIFIED")
    plausible_items = _build_group("plausible", "PLAUSIBLE")
    disputed_items = _build_group("disputed", "DISPUTED")

    subtotal_rows = {}

    def _write_addback_group(ws, row, items, group_label):
        """Write a group of add-back items and a subtotal row. Returns (next_row, subtotal_row)."""
        if not items:
            return row, None

        start_row = row
        for name, info in items:
            ws.cell(row=row, column=1, value=name)
            for i, yr in enumerate(years):
                val = info["amounts"].get(yr, 0)
                cell = ws.cell(row=row, column=year_cols[i], value=val)
                cell.number_format = ACCT_FMT
            ws.cell(row=row, column=rating_col, value=info["rating"])
            ws.cell(row=row, column=rating_col).font = NOTE_FONT
            ws.cell(row=row, column=notes_col, value=info["note"])
            ws.cell(row=row, column=notes_col).font = NOTE_FONT
            thin_border_row(ws, row, 1, max_col)
            row += 1
        end_row = row - 1

        # Subtotal
        ws.cell(row=row, column=1, value=f"Subtotal — {group_label}")
        ws.cell(row=row, column=1).font = TOTAL_FONT
        for i in range(num_years):
            c = year_cols[i]
            cl = get_column_letter(c)
            ws.cell(row=row, column=c).value = f"=SUM({cl}{start_row}:{cl}{end_row})"
            ws.cell(row=row, column=c).number_format = ACCT_FMT
            ws.cell(row=row, column=c).font = TOTAL_FONT
        subtotal_r = row
        row += 1
        return row, subtotal_r

    row += 1
    row, verified_subtotal = _write_addback_group(ws, row, verified_items, "Verified")
    row, plausible_subtotal = _write_addback_group(ws, row, plausible_items, "Plausible")
    row, disputed_subtotal = _write_addback_group(ws, row, disputed_items, "Disputed")

    subtotal_rows["verified"] = verified_subtotal
    subtotal_rows["plausible"] = plausible_subtotal
    subtotal_rows["disputed"] = disputed_subtotal

    # TOTAL ADD-BACKS
    ws.cell(row=row, column=1, value="TOTAL ADD-BACKS")
    total_ab_row = row
    existing_subtotals = [r for r in subtotal_rows.values() if r is not None]
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        parts = [f"{cl}{r}" for r in existing_subtotals]
        ws.cell(row=row, column=c).value = f"={'+'.join(parts)}"
        ws.cell(row=row, column=c).number_format = ACCT_FMT
    style_total_row(ws, row, max_col)

    # ── SELLER'S DISCRETIONARY EARNINGS ──
    row += 2
    ws.cell(row=row, column=1, value="SELLER'S DISCRETIONARY EARNINGS")
    style_section_row(ws, row, max_col)

    # Conservative = Net Income + Verified only
    row += 1
    ws.cell(row=row, column=1, value="Conservative (verified addbacks only)")
    conservative_row = row
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        formula = f"={cl}{ni_row}"
        if verified_subtotal:
            formula += f"+{cl}{verified_subtotal}"
        ws.cell(row=row, column=c).value = formula
        ws.cell(row=row, column=c).number_format = ACCT_FMT_NEG
    ws.cell(row=row, column=notes_col, value="Net Income + verified only")
    ws.cell(row=row, column=notes_col).font = NOTE_FONT

    # Moderate = Net Income + Verified + Plausible
    row += 1
    ws.cell(row=row, column=1, value="Moderate (verified + plausible)")
    moderate_row = row
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        formula = f"={cl}{ni_row}"
        if verified_subtotal:
            formula += f"+{cl}{verified_subtotal}"
        if plausible_subtotal:
            formula += f"+{cl}{plausible_subtotal}"
        ws.cell(row=row, column=c).value = formula
        ws.cell(row=row, column=c).number_format = ACCT_FMT_NEG
    ws.cell(row=row, column=notes_col, value="Net Income + verified + plausible")
    ws.cell(row=row, column=notes_col).font = NOTE_FONT

    # Aggressive = Net Income + All Add-backs
    row += 1
    ws.cell(row=row, column=1, value="Aggressive / Seller Claimed (all addbacks)")
    aggressive_row = row
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f"={cl}{ni_row}+{cl}{total_ab_row}"
        ws.cell(row=row, column=c).number_format = ACCT_FMT_NEG
    if num_years >= 2:
        prev_cl = get_column_letter(year_cols[-2])
        last_cl = get_column_letter(year_cols[-1])
        ws.cell(row=row, column=yoy_col).value = f"=({last_cl}{row}-{prev_cl}{row})/{prev_cl}{row}"
        ws.cell(row=row, column=yoy_col).number_format = PCT_FMT
    ws.cell(row=row, column=notes_col, value="Net Income + all addbacks")
    ws.cell(row=row, column=notes_col).font = NOTE_FONT

    # SDE Margin (Aggressive)
    row += 1
    ws.cell(row=row, column=1, value="  SDE Margin (Aggressive)")
    ws.cell(row=row, column=1).font = FAINT_FONT
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f"={cl}{aggressive_row}/{cl}{rev_row}"
        ws.cell(row=row, column=c).number_format = PCT_FMT

    # ── KEY METRICS ──
    row += 2
    ws.cell(row=row, column=1, value="KEY METRICS")
    style_section_row(ws, row, max_col)

    row += 1
    ws.cell(row=row, column=1, value="Revenue Growth (YoY)")
    if num_years >= 2:
        ws.cell(row=row, column=year_cols[0]).value = "—"
        for i in range(1, num_years):
            c = year_cols[i]
            cl = get_column_letter(c)
            prev_cl = get_column_letter(year_cols[i - 1])
            ws.cell(row=row, column=c).value = f"=({cl}{rev_row}-{prev_cl}{rev_row})/{prev_cl}{rev_row}"
            ws.cell(row=row, column=c).number_format = PCT_FMT

    if has_cogs:
        row += 1
        ws.cell(row=row, column=1, value="COGS as % of Revenue")
        for i in range(num_years):
            c = year_cols[i]
            cl = get_column_letter(c)
            ws.cell(row=row, column=c).value = f"={cl}{cogs_row}/{cl}{rev_row}"
            ws.cell(row=row, column=c).number_format = PCT_FMT

    row += 1
    ws.cell(row=row, column=1, value="OpEx as % of Revenue")
    for i in range(num_years):
        c = year_cols[i]
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f"={cl}{total_exp_row}/{cl}{rev_row}"
        ws.cell(row=row, column=c).number_format = PCT_FMT

    # Freeze panes
    ws.freeze_panes = "B5"

    return {
        "moderate_row": moderate_row,
        "rev_row": rev_row,
        "ni_row": ni_row,
        "years": years,
        "year_cols": year_cols,
    }


def main():
    parser = argparse.ArgumentParser(description="Generate P&L + SDE workbook from deal_state.json")
    parser.add_argument("--deal", required=True, help="Path to deal folder (e.g. deals/my-deal)")
    parser.add_argument("--output", "-o", default=None, help="Output file path (default: deal folder)")
    args = parser.parse_args()

    deal_dir = Path(args.deal)
    if not deal_dir.is_dir():
        print(f"Error: {deal_dir} is not a directory", file=sys.stderr)
        sys.exit(1)

    ds_path = deal_dir / "deal_state.json"
    if not ds_path.exists():
        print(f"Error: {ds_path} not found", file=sys.stderr)
        sys.exit(1)

    data = load_deal_data(deal_dir)
    deal_name = data["deal_name"]

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = deal_dir / f"{deal_name} - Financial Model.xlsx"

    wb = Workbook()
    refs = build_pnl_sheet(wb, data)
    wb.save(str(output_path))
    print(f"P&L + SDE workbook saved to: {output_path}")
    print(f"Moderate SDE row: {refs['moderate_row']}")


if __name__ == "__main__":
    main()
